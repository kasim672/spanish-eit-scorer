"""
score_test_data.py
====================
GSoC 2026 AutoEIT — Test II: Evaluation of transcribed data

Applies the meaning-based EIT rubric to the sample transcriptions
(participant 38006-2A) and outputs sentence-level scores to an Excel file.

Usage:
    python score_test_data.py
    python score_test_data.py --participant P1 --out results/scored_P1.xlsx
    python score_test_data.py --all --out results/all_scored.xlsx

Output columns (matching AutoEIT Sample Transcriptions for Scoring.xlsx):
    Sentence  | Stimulus | Transcription Rater 1 | Score | Rule | Edits | Overlap | Similarity

Approach:
    1. Normalize transcriptions (strip disfluencies: [pause], [gibberish],
       xxx, partial words like "ma-", filler sounds)
    2. Tokenize with Spanish enclitic splitting
    3. Align learner response to target using Needleman-Wunsch
    4. Count error types (substitutions, omissions, insertions)
    5. Apply meaning-based rubric (YAML-configurable, first-match-wins)
    6. Compute PyTorch similarity score as a continuous complement
    7. Write results to Excel
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

# ── Make importable from project root ────────────────────────
sys.path.insert(0, str(Path(__file__).parent))

from eit_scorer.core.rubric      import load_rubric
from eit_scorer.core.scoring     import score_response
from eit_scorer.core.similarity  import compute_similarity
from eit_scorer.data.models      import EITItem, EITResponse

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    _XLSX = True
except ImportError:
    _XLSX = False


# ─────────────────────────────────────────────────────────────
# PARTICIPANT DATA  (from AutoEIT Sample Transcriptions for Scoring.xlsx)
# Tab: 38006-2A  — one participant, 30 sentences
# Format: (sentence_num, target_sentence, transcription)
# ─────────────────────────────────────────────────────────────

PARTICIPANT_DATA = {
    "38006-2A": [
        (1,  "Quiero cortarme el pelo",                                          "Quiero cortarme mi pelo"),
        (2,  "El libro está en la mesa",                                          "El libro [pause] está en la mesa"),
        (3,  "El carro lo tiene Pedro",                                           "E-[gibberish] perro"),
        (4,  "El se ducha cada mañana",                                           "El se lucha cada mañana"),
        (5,  "¿Qué dice usted que va a hacer hoy?",                               "¿Qué [gibberish] que vas estoy?"),
        (6,  "Dudo que sepa manejar muy bien",                                    "Dudo que sepa ma-mastar tan bien (tambien?)"),
        (7,  "Las calles de esta ciudad son muy anchas",                          "Las calles..es-[gibberish]..."),
        (8,  "Puede que llueva mañana todo el día",                               "Puede xxx mañana de todo día"),
        (9,  "Las casas son muy bonitas pero caras",                              "A las casa es mu-son bonitas"),
        (10, "Me gustan las películas que acaban bien",                           "Me gusta las películas que x bien"),
        (11, "El chico con el que yo salgo es español",                           "El chico con es el algo (xxx?) es español"),
        (12, "Después de cenar me fui a dormir tranquilo",                        "Después de fenar mi fui a tranquilo"),
        (13, "Quiero una casa en la que vivan mis animales",                      "Quienen que en el casa ...muchos animales"),
        (14, "A nosotros nos fascinan las fiestas grandiosas",                    "A nosotros fanimos [pause] fiestas grandiosas"),
        (15, "Ella sólo bebe cerveza y no come nada",                             "Ella bebidas cerviamos y no comidas"),
        (16, "Me gustaría que el precio de las casas bajara",                     "Me gustaría ser..ella...mhh.."),
        (17, "Cruza a la derecha y después sigue todo recto",                     "Cruze a la derecha de siguelo"),
        (18, "Ella ha terminado de pintar su apartamento",                        "Ella terminado pintar a sus apartmento"),
        (19, "Me gustaría que empezara a hacer más calor pronto",                 "Me gustaría se..[pause] xxx pronto"),
        (20, "El niño al que se le murió el gato está triste",                    "El niño se murió el gato es muy triste"),
        (21, "Una amiga mía cuida a los niños de mi vecino",                      "Una mía amiga cuidado a sus niños"),
        (22, "El gato que era negro fue perseguido por el perro",                 "El gato-el gato quien era el nego de perro"),
        (23, "Antes de poder salir él tiene que limpiar su cuarto",               "Antes de podai e salir..[pause] antes de xxx"),
        (24, "La cantidad de personas que fuman ha disminuido",                   "A la cantan... [pause] muy a xxx"),
        (25, "Después de llegar a casa del trabajo tomé la cena",                 "Después de llegar [pause] al trabajo..."),
        (26, "El ladrón al que atrapó la policía era famoso",                     "X- ladróna(?)de policio que es famoso"),
        (27, "Le pedí a un amigo que me ayudara con la tarea",                    "Que ped-x una amigo [pause] a su tarea"),
        (28, "El examen no fue tan difícil como me habían dicho",                 "El examen no difícil.. [pause]... [gibberish]"),
        (29, "¿Serías tan amable de darme el libro que está en la mesa?",         "El libro que está en la mesa"),
        (30, "Hay mucha gente que no toma nada para el desayuno",                 "A mucha gente xxx..."),
    ]
}


# ─────────────────────────────────────────────────────────────
# TRANSCRIPTION CLEANING
# Strips disfluency markers before scoring.
# These are transcription annotations, not learner words.
# ─────────────────────────────────────────────────────────────

_DISFLUENCY_RE = re.compile(
    r'\[pause\]|\[gibberish\]|\[.*?\]'   # bracketed annotations
    r'|xxx?'                              # unintelligible markers
    r'|\b\w+-\b'                          # partial words: "ma-", "E-"
    r'|\(.*?\)'                           # parenthetical notes: (tambien?)
    r'|\.{2,}'                            # ellipsis
    r'|[¿¡]',                             # inverted punctuation
    re.IGNORECASE
)


def clean_transcription(text: str) -> str:
    """Remove disfluency markers, keeping only scoreable content."""
    cleaned = _DISFLUENCY_RE.sub(' ', text)
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    return cleaned


# ─────────────────────────────────────────────────────────────
# SCORING
# ─────────────────────────────────────────────────────────────

def score_participant(
    participant_id: str,
    data: list[tuple],
    rubric,
) -> list[dict]:
    """Score all 30 sentences for one participant."""
    results = []

    for sent_num, target, raw_transcription in data:
        cleaned = clean_transcription(raw_transcription)

        item = EITItem(
            item_id=f"s{sent_num:02d}",
            reference=target,
            max_points=4,
        )
        resp = EITResponse(
            participant_id=participant_id,
            item_id=f"s{sent_num:02d}",
            response_text=cleaned,
        )

        ss = score_response(item, resp, rubric)

        # PyTorch similarity score (continuous complement to rule score)
        sim = compute_similarity(
            overlap_ratio=ss.trace.overlap_ratio,
            total_edits=ss.trace.total_edits,
            content_subs=ss.trace.content_subs,
            ref_len=len(ss.trace.ref_tokens),
            hyp_len=len(ss.trace.hyp_tokens),
        )

        results.append({
            "sentence":        sent_num,
            "stimulus":        target,
            "transcription":   raw_transcription,
            "cleaned":         cleaned,
            "score":           ss.score,
            "max_points":      ss.max_points,
            "rule":            ss.trace.applied_rule_ids[0] if ss.trace.applied_rule_ids else "",
            "total_edits":     ss.trace.total_edits,
            "overlap_ratio":   round(ss.trace.overlap_ratio, 3),
            "similarity":      round(sim, 3),
            "ref_tokens":      ss.trace.ref_tokens,
            "hyp_tokens":      ss.trace.hyp_tokens,
        })

    return results


# ─────────────────────────────────────────────────────────────
# EXCEL OUTPUT
# ─────────────────────────────────────────────────────────────

_SCORE_COLORS = {
    4: "C6EFCE",  # green
    3: "FFEB9C",  # yellow
    2: "FFCC99",  # orange
    1: "FFB3B3",  # light red
    0: "FF6666",  # red
}


def write_excel(results: list[dict], participant_id: str, out_path: Path) -> None:
    """Write scored results to Excel matching the AutoEIT format."""
    if not _XLSX:
        print("openpyxl not installed — writing CSV instead")
        _write_csv(results, out_path.with_suffix(".csv"))
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = participant_id

    # ── Header ──
    headers = [
        "Sentence", "Stimulus", "Transcription Rater 1",
        "Score", "Rule Fired", "Total Edits", "Overlap", "Similarity (PyTorch)"
    ]
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border

    # ── Data rows ──
    for row_idx, r in enumerate(results, 2):
        values = [
            r["sentence"],
            r["stimulus"],
            r["transcription"],
            r["score"],
            r["rule"],
            r["total_edits"],
            r["overlap_ratio"],
            r["similarity"],
        ]
        score_color = _SCORE_COLORS.get(r["score"], "FFFFFF")

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 4:  # Score column — color by value
                cell.fill = PatternFill("solid", fgColor=score_color)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

    # ── Column widths ──
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 20

    # ── Summary row ──
    total = sum(r["score"] for r in results)
    max_total = sum(r["max_points"] for r in results)
    summary_row = len(results) + 2
    ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
    score_cell = ws.cell(row=summary_row, column=4, value=f"{total}/{max_total}")
    score_cell.font = Font(bold=True)
    score_cell.alignment = Alignment(horizontal="center")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


def _write_csv(results: list[dict], out_path: Path) -> None:
    import csv
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Sentence", "Stimulus", "Transcription", "Score",
                    "Rule", "Edits", "Overlap", "Similarity"])
        for r in results:
            w.writerow([r["sentence"], r["stimulus"], r["transcription"],
                        r["score"], r["rule"], r["total_edits"],
                        r["overlap_ratio"], r["similarity"]])
    print(f"Saved: {out_path}")


# ─────────────────────────────────────────────────────────────
# CONSOLE REPORT
# ─────────────────────────────────────────────────────────────

def print_report(results: list[dict], participant_id: str) -> None:
    total = sum(r["score"] for r in results)
    max_total = sum(r["max_points"] for r in results)

    print(f"\n{'='*72}")
    print(f"  AutoEIT Test II — Scoring Report")
    print(f"  Participant: {participant_id}")
    print(f"{'='*72}")
    print(f"  {'#':>2}  {'Score':>5}  {'Rule':<28}  {'Ed':>3}  {'Ovlp':>5}  {'Sim':>5}")
    print(f"  {'-'*62}")

    for r in results:
        print(f"  {r['sentence']:>2}  {r['score']:>5}  {r['rule']:<28}  "
              f"{r['total_edits']:>3}  {r['overlap_ratio']:>5.2f}  {r['similarity']:>5.3f}")

    print(f"  {'-'*62}")
    print(f"  {'TOTAL':>2}  {total:>5}/{max_total}  ({total/max_total*100:.1f}%)")
    print(f"{'='*72}\n")

    # Score distribution
    from collections import Counter
    dist = Counter(r["score"] for r in results)
    print("  Score distribution:")
    for s in sorted(dist, reverse=True):
        bar = "█" * dist[s]
        print(f"    {s}/4 : {bar} ({dist[s]})")
    print()


# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        prog="score_test_data",
        description="GSoC AutoEIT Test II — Score sample transcriptions",
    )
    parser.add_argument("--participant", default="38006-2A",
                        choices=list(PARTICIPANT_DATA.keys()),
                        help="Participant ID to score (default: 38006-2A)")
    parser.add_argument("--out", default="results/AutoEIT_Scored_38006-2A.xlsx",
                        help="Output Excel file path")
    parser.add_argument("--rubric", default=None,
                        help="Path to rubric YAML (default: built-in)")
    parser.add_argument("--no-excel", action="store_true",
                        help="Print report only, no file output")
    args = parser.parse_args()

    rubric_path = args.rubric or (
        Path(__file__).parent / "eit_scorer" / "config" / "default_rubric.yaml"
    )
    rubric = load_rubric(rubric_path)

    print(f"\nRubric : {rubric.summary()}")
    print(f"Participant: {args.participant}")

    data    = PARTICIPANT_DATA[args.participant]
    results = score_participant(args.participant, data, rubric)

    print_report(results, args.participant)

    if not args.no_excel:
        write_excel(results, args.participant, Path(args.out))


if __name__ == "__main__":
    main()
