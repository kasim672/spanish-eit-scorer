"""
eit_scorer/utils/excel_io.py
=============================
Excel file I/O for EIT scoring.

Handles reading and writing Excel files with multiple sheets.
Skips "Info" sheet and processes all other sheets as participant datasets.
"""

from __future__ import annotations

from pathlib import Path
from dataclasses import dataclass

import openpyxl


@dataclass
class ExcelRow:
    """A single row from an Excel sheet."""
    sentence_id: int | str
    stimulus: str
    transcription: str
    score: int | None = None


def read_excel_sheets(input_path: str | Path) -> dict[str, list[ExcelRow]]:
    """
    Read all sheets from an Excel file (except 'Info').
    
    Returns:
        dict mapping sheet_name → list of ExcelRow objects
    """
    input_path = Path(input_path)
    if not input_path.exists():
        raise FileNotFoundError(f"Excel file not found: {input_path}")
    
    wb = openpyxl.load_workbook(input_path)
    sheets_data: dict[str, list[ExcelRow]] = {}
    
    for sheet_name in wb.sheetnames:
        # Skip Info sheet
        if sheet_name.lower() == "info":
            continue
        
        ws = wb[sheet_name]
        rows: list[ExcelRow] = []
        
        # Skip header row (row 1)
        for row_idx in range(2, ws.max_row + 1):
            sentence_id = ws.cell(row_idx, 1).value
            stimulus = ws.cell(row_idx, 2).value
            transcription = ws.cell(row_idx, 3).value
            score = ws.cell(row_idx, 4).value
            
            # Skip empty rows
            if not stimulus or not transcription:
                continue
            
            rows.append(ExcelRow(
                sentence_id=sentence_id,
                stimulus=str(stimulus).strip(),
                transcription=str(transcription).strip(),
                score=score,
            ))
        
        sheets_data[sheet_name] = rows
    
    return sheets_data


def write_excel_scores(
    input_path: str | Path,
    output_path: str | Path,
    scores_by_sheet: dict[str, list[int]],
) -> None:
    """
    Write scores back to Excel file.
    
    Args:
        input_path: Path to input Excel file
        output_path: Path to output Excel file
        scores_by_sheet: dict mapping sheet_name → list of scores (in row order)
    """
    input_path = Path(input_path)
    output_path = Path(output_path)
    
    wb = openpyxl.load_workbook(input_path)
    
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() == "info":
            continue
        
        if sheet_name not in scores_by_sheet:
            continue
        
        ws = wb[sheet_name]
        scores = scores_by_sheet[sheet_name]
        
        # Write scores starting from row 2 (skip header)
        score_idx = 0
        for row_idx in range(2, ws.max_row + 1):
            stimulus = ws.cell(row_idx, 2).value
            transcription = ws.cell(row_idx, 3).value
            
            # Skip empty rows
            if not stimulus or not transcription:
                continue
            
            if score_idx < len(scores):
                ws.cell(row_idx, 4).value = scores[score_idx]
                score_idx += 1
    
    wb.save(output_path)
